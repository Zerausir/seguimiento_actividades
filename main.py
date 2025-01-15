import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, NamedStyle, Font
import glob
from environs import Env, EnvError

# Inicializar environs
env = Env()
env.read_env()


def get_indicator_path(indicator, year):
    """Retorna la ruta correspondiente al indicador y año"""
    indicator_var = f"RUTA_{indicator}_{year}"
    return env.str(indicator_var, None)


def get_monthly_indicator_path(indicator, month, year):
    """Retorna la ruta mensual para indicadores que la tienen"""
    indicator_var = f"RUTA_{indicator}_{month}_{year}"
    return env.str(indicator_var, None)


def check_file_exists(informe_num, indicator, year):
    """
    Verifica si el archivo existe en las rutas correspondientes
    """
    if not informe_num:
        return False

    # Convertir a string y eliminar espacios
    informe_num = str(informe_num).strip()

    # Primero buscar en el directorio cronológico general
    cronologico_path = env.str(f'RUTA_CRONOLOGICO_{year}', None)
    if cronologico_path and os.path.exists(cronologico_path):
        pattern = os.path.join(cronologico_path, f"{informe_num}*")
        if glob.glob(pattern):
            return True

    # Buscar en la ruta específica del indicador
    indicator_path = get_indicator_path(indicator, year)
    if indicator_path and os.path.exists(indicator_path):
        pattern = os.path.join(indicator_path, f"{informe_num}*")
        if glob.glob(pattern):
            return True

    # Si el indicador es CCDE-01, buscar en las carpetas mensuales
    if indicator == 'CCDE-01':
        for month in ['ENE', 'FEB', 'MAR', 'ABR', 'MAY', 'JUN',
                      'JUL', 'AGO', 'SEP', 'OCT', 'NOV', 'DIC']:
            monthly_path = get_monthly_indicator_path(indicator, month, year)
            if monthly_path and os.path.exists(monthly_path):
                pattern = os.path.join(monthly_path, f"{informe_num}*")
                if glob.glob(pattern):
                    return True

    return False


def verify_environment_variables():
    """Verifica que las variables de entorno necesarias estén disponibles"""
    required_vars = ['SERVER_ROUTE', 'DOWNLOAD_ROUTE']
    missing_vars = []

    for var in required_vars:
        try:
            env.str(var)
        except EnvError:
            missing_vars.append(var)

    if missing_vars:
        raise EnvironmentError(f"Faltan las siguientes variables de entorno: {', '.join(missing_vars)}")


def process_seguimiento(file_path):
    # Verificar variables de entorno
    verify_environment_variables()

    # Define color fills
    green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
    red_fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    # Define border style
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Leer el archivo Excel
    df = pd.read_excel(file_path)

    # Copiar valores originales de ATENDIDO en la nueva columna ATENDIDO FUENTE
    df.insert(df.columns.get_loc('ATENDIDO'), 'ATENDIDO FUENTE', df['ATENDIDO'])

    # Crear columna INDICADOR AÑO ENCONTRADO al lado de INDICADOR AÑO
    df.insert(df.columns.get_loc('INDICADOR AÑO') + 1, 'INDICADOR AÑO ENCONTRADO', '')

    # Actualizar ATENDIDO y llenar INDICADOR AÑO ENCONTRADO
    for index, row in df.iterrows():
        informe = row.get('INFORME')
        indicador = row.get('INDICADOR CORTO')
        year = row.get('INDICADOR AÑO')

        file_found = False
        if pd.notna(informe):  # Si hay un valor en la columna INFORME
            # Verificar si existe en cronologico_path
            for y in ['2023', '2024', '2025']:  # Años a verificar
                cronologico_path = env.str(f'RUTA_CRONOLOGICO_{y}', None)
                if cronologico_path and os.path.exists(cronologico_path):
                    pattern = os.path.join(cronologico_path, f"{str(informe).strip()}*")
                    if glob.glob(pattern):
                        file_found = True
                        df.at[index, 'INDICADOR AÑO ENCONTRADO'] = int(y)  # Guardar el año como entero
                        break

            # Verificar en la ruta específica del indicador
            if not file_found:
                file_found = check_file_exists(informe, indicador, year)

            # Actualizar ATENDIDO si se encontró el archivo
            if file_found:
                df.at[index, 'ATENDIDO'] = 'SI'

    # Guardar archivo actualizado
    download_path = env.str('DOWNLOAD_ROUTE')
    output_file = os.path.join(download_path, env.str('RESULTS'))
    df.to_excel(output_file, index=False)

    # Aplicar formato condicional
    wb = load_workbook(output_file)
    ws = wb.active

    # Obtener índices de las columnas
    def get_column_index(column_name):
        for idx, cell in enumerate(ws[1], 1):
            if cell.value == column_name:
                return idx
        return None

    col_indices = {
        'ATENDIDO': get_column_index('ATENDIDO'),
        'ATENDIDO FUENTE': get_column_index('ATENDIDO FUENTE'),
        'INFORME': get_column_index('INFORME'),
        'FECHA_INF': get_column_index('FECHA_INF'),
        'PARA_JURIDICO': get_column_index('PARA JURÍDICO'),
        'MEMO_JUR': get_column_index('ENVIADO A JUR CON MEMORANDO Nro.'),
        'FECHA_JUR': get_column_index('FECHA_JUR'),
        'MEMO_DT': get_column_index('ENVIADO POR DT CON MEMORANDO Nro.'),
        'FECHA_DT': get_column_index('FECHA_DT'),
        'INDICADOR': get_column_index('INDICADOR CORTO'),
        'INDICADOR AÑO': get_column_index('INDICADOR AÑO'),
        'INDICADOR AÑO ENCONTRADO': get_column_index('INDICADOR AÑO ENCONTRADO')
    }

    # Aplicar reglas de formato condicional
    for row in range(2, ws.max_row + 1):
        atendido = ws.cell(row=row, column=col_indices['ATENDIDO']).value
        atendido_fuente = ws.cell(row=row, column=col_indices['ATENDIDO FUENTE']).value
        informe = ws.cell(row=row, column=col_indices['INFORME']).value
        para_juridico = ws.cell(row=row, column=col_indices['PARA_JURIDICO']).value
        memo_jur = ws.cell(row=row, column=col_indices['MEMO_JUR']).value
        memo_dt = ws.cell(row=row, column=col_indices['MEMO_DT']).value
        year = ws.cell(row=row, column=col_indices['INDICADOR AÑO']).value
        year_found = ws.cell(row=row, column=col_indices['INDICADOR AÑO ENCONTRADO']).value

        # Rellenar en amarillo si los años son diferentes
        if year_found and year_found != year:
            ws.cell(row=row, column=col_indices['INDICADOR AÑO']).fill = yellow_fill
            ws.cell(row=row, column=col_indices['INDICADOR AÑO ENCONTRADO']).fill = yellow_fill
        elif year_found and year_found == year:
            ws.cell(row=row, column=col_indices['INDICADOR AÑO']).fill = green_fill
            ws.cell(row=row, column=col_indices['INDICADOR AÑO ENCONTRADO']).fill = green_fill

        # Rellenar en rojo si no hay valor en INDICADOR AÑO ENCONTRADO
        if not year_found:
            ws.cell(row=row, column=col_indices['INDICADOR AÑO ENCONTRADO']).fill = red_fill

        # Rellenar en amarillo si ATENDIDO FUENTE es diferente
        if atendido_fuente != atendido:
            ws.cell(row=row, column=col_indices['ATENDIDO FUENTE']).fill = yellow_fill

        # Regla 1: Colorear ATENDIDO
        ws.cell(row=row, column=col_indices['ATENDIDO']).fill = green_fill if atendido == 'SI' else red_fill

        # Regla 2: Colorear INFORME y FECHA_INF
        if year_found:
            ws.cell(row=row, column=col_indices['INFORME']).fill = green_fill
            ws.cell(row=row, column=col_indices['FECHA_INF']).fill = green_fill
        else:
            ws.cell(row=row, column=col_indices['INFORME']).fill = red_fill
            ws.cell(row=row, column=col_indices['FECHA_INF']).fill = red_fill

        # Regla 3: Colorear columnas de JURÍDICO
        if atendido == 'SI':
            if para_juridico in ['AP', 'FI']:
                if memo_jur:
                    for col in ['PARA_JURIDICO', 'MEMO_JUR', 'FECHA_JUR']:
                        ws.cell(row=row, column=col_indices[col]).fill = green_fill
                else:
                    for col in ['PARA_JURIDICO', 'MEMO_JUR', 'FECHA_JUR']:
                        ws.cell(row=row, column=col_indices[col]).fill = red_fill
            elif not para_juridico or para_juridico == 'NO':
                for col in ['PARA_JURIDICO', 'MEMO_JUR', 'FECHA_JUR']:
                    ws.cell(row=row, column=col_indices[col]).fill = green_fill

        # Regla 4: Colorear columnas de DT
        if atendido == 'SI':
            fill_color = green_fill if memo_dt else red_fill
            ws.cell(row=row, column=col_indices['MEMO_DT']).fill = fill_color
            ws.cell(row=row, column=col_indices['FECHA_DT']).fill = fill_color

        # Regla 5: Todo rojo si ATENDIDO es NO y no se cumple ninguna condición
        if atendido == 'NO':
            if not informe and not memo_jur and not memo_dt:
                for col in ['ATENDIDO', 'INFORME', 'FECHA_INF', 'PARA_JURIDICO',
                            'MEMO_JUR', 'FECHA_JUR', 'MEMO_DT', 'FECHA_DT']:
                    ws.cell(row=row, column=col_indices[col]).fill = red_fill
            elif informe and not year_found:
                for col in ['ATENDIDO', 'INFORME', 'FECHA_INF', 'PARA_JURIDICO', 'MEMO_JUR', 'FECHA_JUR', 'MEMO_DT',
                            'FECHA_DT']:
                    ws.cell(row=row, column=col_indices[col]).fill = red_fill

    # Ajustar el ancho de las columnas automáticamente
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 25)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Habilitar filtros y congelar la primera fila
    ws.auto_filter.ref = f"A1:{ws.cell(row=1, column=ws.max_column).column_letter}1"
    ws.freeze_panes = "G2"

    # Guardar el archivo actualizado
    wb.save(output_file)
    print(f"Archivo consolidado guardado como: {output_file}")


# Execute the processing
if __name__ == "__main__":
    input_file = os.path.join(env.str('SERVER_ROUTE'), env.str('FILE_SEGUIMIENTO_GPR'))
    process_seguimiento(input_file)
